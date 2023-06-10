import datetime

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame


def create_and_save_excel(data, config):
    frame = []
    # calculate the real hours to work per day from the config file
    full_time_hours = int(config["full_time_hours"])
    daily_percentage = float(config["daily_percentage"]) / 100
    target_time = int(full_time_hours * daily_percentage)
    for year, months in data.items():
        for month, days in months.items():
            for day, details in days.items():
                date = datetime.datetime(year=int(year), month=int(month), day=int(day))
                date_str = date.strftime('%Y-%m-%d')
                day_of_week = date.strftime('%A')
                start_time = details['start_job']
                start_lunch_break = details['start_break']
                end_lunch_break = details['end_break']
                end_time = details['end_job']
                location = details['location']
                overtime = details['under_or_over']
                hours_diff = details['time_diff']
                frame.append(
                    {
                        "date": date_str,
                        "date_dt": date,
                        "day_of_week": day_of_week,
                        "start_time": start_time,
                        "start_lunch_break": start_lunch_break,
                        "end_lunch_break": end_lunch_break,
                        "end_time": end_time,
                        "location": location,
                        "overtime": overtime,
                        "hours_diff": hours_diff
                    }
                )

    df = DataFrame(frame)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Create a new column for month-year
    df['month_year'] = df['date_dt'].dt.strftime('%B %Y')

    # Drop the 'date_dt' column
    df = df.drop(columns=['date_dt'])

    # Now iterate over the unique 'month_year' instead of 'date_dt'
    for month_year_name in df['month_year'].unique():
        df_month = df[df['month_year'] == month_year_name]

        if not df_month.empty:
            ws = wb.create_sheet(title=str(month_year_name))

            for r in dataframe_to_rows(df_month, index=False, header=True):
                ws.append(r)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
                for cell in row:
                    if cell.value is not None:
                        value_str = str(cell.value).strip('-')
                        if ':' in value_str:
                            hours, minutes = value_str.split(':')
                            value = float(hours) + float(minutes) / 60.0
                        elif value_str.replace('.', '', 1).isdigit() or value_str.replace('.', '', 1).lstrip('-').isdigit():
                            value = float(value_str)
                        else:
                            value = None

                        if str(cell.value).startswith('-'):
                            value = -value

                        if value is not None:
                            if value < 0:
                                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            elif value > 0:
                                fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
                            else:
                                fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        else:
                            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                        cell.fill = fill

                    day_row = cell.row
                    day_of_week = df_month.iloc[day_row - 2]["day_of_week"]
                    if day_of_week in ["Saturday", "Sunday"] and cell.value is not None and cell.value != 0:
                        for cell_in_row in ws[day_row]:
                            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            cell_in_row.fill = fill

            total_minutes = 0
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
                for cell in row:
                    if cell.value is not None:
                        value_str = str(cell.value)
                        print(value_str)
                        is_negative = value_str.startswith('-')
                        value_str = value_str.strip('-')
                        if ':' in value_str:
                            hours, minutes = map(int, value_str.split(':'))
                            value = hours * 60 + minutes
                            total_minutes += -value if is_negative else value
                            print(total_minutes)

            total_hours, total_minutes = divmod(abs(total_minutes), 60)
            sign = "-" if total_minutes < 0 else ""
            total_hours_diff_str = '{}{}:{:02d}'.format(sign, int(total_hours), int(total_minutes))

            ws.append([""] * 11 + ["Total hours diff for the month", total_hours_diff_str])

    wb.save("hours_work.xlsx")
