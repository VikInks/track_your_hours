import datetime
import json
import os
import signal
import sys

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

from utils.calculate_work_hours_and_difference import calculate_work_hours_and_difference
from utils.prompt import prompt_for_month, prompt_for_day, prompt_for_year
from utils.validate_input import validate_input


class WorkHoursTracker:
    def __init__(self):
        try:
            with open('work_hours.json', 'r') as f:
                self.work_hours_data = json.load(f)
        except FileNotFoundError:
            self.work_hours_data = []

            # Print an error message or handle the exception appropriately
            print("work_hours.json file not found. Starting with an empty data list.")

    def recalculate_work_hours_difference(self):
        print("Recalculating work hours and difference...")
        """Recalculate the work hours and difference for all entries."""
        for entry in self.work_hours_data:
            # Calculate the work hours and difference
            updated_data_dict = calculate_work_hours_and_difference(
                entry["date"],
                entry["start_time"],
                datetime.timedelta(hours=int(entry["lunch_break"].split(":")[0]),
                                   minutes=int(entry["lunch_break"].split(":")[1])),
                entry["end_time"],
                entry["location"]
            )
            # Update the entry with the updated data
            entry.update(updated_data_dict)

    def signal_handler(self, sig, frame):
        """Handle Ctrl+C interruption and save current data."""
        print('Interruption detected, saving data...')
        self.save_to_json()
        sys.exit(0)

    def save_to_json(self):
        """Save data to a JSON file."""
        # Ensure all keys are present in each dictionary in work_hours_data
        for entry in self.work_hours_data:
            for key in ["date", "day_of_week", "start_time", "lunch_break", "end_time", "location", "overtime",
                        "work_hours", "hours_diff"]:
                entry.setdefault(key, "")

        # Create a temporary file to avoid data loss in case of a crash during the write operation
        tempname = os.path.join(os.path.dirname('work_hours.json'), "tmp_" + os.path.basename('work_hours.json'))
        with open(tempname, 'w') as json_file:
            json.dump(self.work_hours_data, json_file, default=str)

        # Remove existing file if it exists
        if os.path.exists('work_hours.json'):
            os.remove('work_hours.json')

        # Rename temporary file to the final name
        os.rename(tempname, 'work_hours.json')

    def create_and_save_excel(self):
        # Create a pandas DataFrame from the work day details
        df = pd.DataFrame(self.work_hours_data,
                          columns=["date", "day_of_week", "start_time", "lunch_break", "end_time", "location",
                                   "overtime", "work_hours", "hours_diff"])

        # Format the 'date' column to datetime dd mm yyyy
        df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d')

        # Get the month name and year from the 'date' column
        month_year_name = df['date'].dt.strftime('%B %Y').iloc[0]

        # Save the DataFrame to an Excel file with the sheet name as the month name
        df.to_excel("work_hours.xlsx", index=False, sheet_name=str(month_year_name))

        # Load the Excel file
        wb = openpyxl.load_workbook("work_hours.xlsx")
        ws = wb[str(month_year_name)]

        # Define the yellow fill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Define a differential style with the yellow fill
        diff_style_yellow = DifferentialStyle(fill=yellow_fill)

        # Define a rule with the style and the formula
        rule_weekend = Rule(type="expression", dxf=diff_style_yellow)
        rule_weekend.formula = ["WEEKDAY($A1, 2)>5"]

        # Add the rule to the conditional formatting of the worksheet
        ws.conditional_formatting.add("A1:I" + str(len(df) + 1), rule_weekend)

        # Define the green and red fill
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Define a differential style with the green and red fill
        diff_style_green = DifferentialStyle(fill=green_fill)
        diff_style_red = DifferentialStyle(fill=red_fill)

        # Define a rule with the style and the formula for positive hours_diff
        rule_positive = Rule(type="expression", dxf=diff_style_green)
        rule_positive.formula = ["LEFT($I1, 1) = '+'"]

        # Define a rule with the style and the formula for negative hours_diff
        rule_negative = Rule(type="expression", dxf=diff_style_red)
        rule_negative.formula = ["LEFT($I1, 1) = '-'"]

        # Add the rule to the conditional formatting of the worksheet
        ws.conditional_formatting.add("I2:I" + str(len(df) + 1), rule_positive)
        ws.conditional_formatting.add("I2:I" + str(len(df) + 1), rule_negative)

        # Save the Excel file
        wb.save("work_hours.xlsx")

    def update_date_work_hours(self, selected_day: str, selected_month: str, selected_year: str):
        """Update the work hours for the given date."""
        if selected_month.__len__() == 1:
            selected_month = f"0{selected_month}"
        if selected_day.__len__() == 1:
            selected_day = f"0{selected_day}"

        date = f"{selected_year}-{selected_month}-{selected_day}"
        entry = next((item for item in self.work_hours_data if item["date"] == date), None)
        if entry is None:
            print("No data found for the selected date.")
            exit()
        # use validate_input() to get the updated data
        updated_data = validate_input(entry)
        # calculate the work hours and difference
        updated_data_dict = calculate_work_hours_and_difference(*updated_data)
        # update the entry with the updated data
        entry.update(updated_data_dict)

        if entry is None:
            print("No data found for the selected date.")
            return

    def run(self):
        """Main function to run the program."""
        signal.signal(signal.SIGINT, self.signal_handler)

        # Check for any incomplete entries
        incomplete_entry = next((entry for entry in self.work_hours_data if entry.get('status') == 'incomplete'), None)
        if incomplete_entry:
            print(f"Incomplete data detected for date {incomplete_entry['date']}. Please enter the missing data.")
            updated_data = validate_input(incomplete_entry)
            updated_data_dict = calculate_work_hours_and_difference(*updated_data)
            incomplete_entry.update(updated_data_dict)
            incomplete_entry['status'] = 'complete'  # Mark the entry as complete

        self.recalculate_work_hours_difference()
        while True:
            # clear the console screen
            os.system('cls' if os.name == 'nt' else 'clear')
            print("\n1. Enter work hours")
            print("2. Modify work hours for a date")
            print("3. Exit")
            choice = input("Enter your choice: ")

            if choice == "1":
                data = validate_input()
                data_dict = calculate_work_hours_and_difference(*data)
                if "day_of_week" not in data_dict:
                    date_datetime = datetime.datetime.strptime(data_dict["date"], "%Y-%m-%d")
                    data_dict["day_of_week"] = date_datetime.strftime("%A")
                data_dict['status'] = 'incomplete'  # Mark the entry as incomplete
                self.work_hours_data.append(data_dict)
                self.save_to_json()  # Save the incomplete entry
                data_dict['status'] = 'complete'  # Mark the entry as complete after saving
            elif choice == '2':
                selected_year = prompt_for_year()
                print("You selected: ", selected_year)
                selected_month = prompt_for_month()
                print("You selected: ", selected_month)
                selected_day = prompt_for_day()
                print("You selected: ", selected_day)
                self.update_date_work_hours(selected_day, selected_month, selected_year)
                self.save_to_json()
            elif choice == "3":
                self.save_to_json()
                self.create_and_save_excel()
                break
            else:
                print("Invalid choice. Please try again.")


if __name__ == "__main__":
    tracker = WorkHoursTracker()
    tracker.run()
